"""Part 3 / Round 3: leakage-free preprocessing for JSFD001--JSFD014.

This module is deliberately independent from TensorFlow.  It rebuilds every
artifact from the original Excel workbooks and refuses to consume the
``processed_npz`` directory created on another branch.

The generated contract is shared by the Round-3 training and prediction
entrypoints:

* ``prepared_data/feature_arrays/<farm>.npz`` stores the scaled full timeline
  and the exact train/validation/test forecast origins.
* ``preprocess/<farm>/preprocessing_bundle.joblib`` stores immutable metadata,
  scaler states, the train-only power reference and WindPRISM regime config.
* ``round3_preprocess_bundle_complete.json`` is written only for a complete
  fourteen-farm formal run.

All timestamps are converted to an *availability-time* axis.  The source files
do not document whether interval statistics are start- or end-labelled, so the
formal default is conservative: raw timestamps are treated as interval starts.
Use ``--timestamp-semantics`` only when authoritative station metadata exists.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.preprocessing import StandardScaler


PROTOCOL_VERSION = "part3_round3_external14_leakage_free_v1"
SCHEMA_VERSION = "FEATURE_SCHEMA_V1"
ARTIFACT_SCHEMA_VERSION = 1
# Verified once against the original five-farm F7 artifact ``input_cols``.
LEGACY_F7_SCHEMA_HASH = (
    "a2f44e932044c2609a8c0e1cf6a446f37b4a0cfb71b8bf232a5bae6c568c680c"
)

RAW_ROOT = Path("./wind_split/supplementary_other_wind_data")
RESULT_ROOT = Path(
    "./wind_results/part3_new_module_supplement/"
    "03_external14_leakage_free_strong_baseline_benchmark"
)
EXPECTED_FARMS = tuple(f"JSFD{i:03d}" for i in range(1, 15))

TIME_FREQ = "15min"
HISTORY_LEN = 96
FORECAST_LEN = 16
TARGET_COL = "功率"
RANDOM_SEED = 2026
CAUSAL_FILL_LIMIT = 4
POWER_REFERENCE_QUANTILE = 0.999
BLANK_ROW_STOP = 512

SPEED_COLS = (
    "10米风速",
    "30米风速",
    "50米风速",
    "70米风速",
    "轮毂高度风速",
)
DIRECTION_BASE_COLS = (
    "10米风向",
    "30米风向",
    "50米风向",
    "70米风向",
    "轮毂高度风向",
)
MET_COLS = ("10m气温", "10m气压", "10m湿度")
HEIGHTS = {
    "10米风速": 10.0,
    "30米风速": 30.0,
    "50米风速": 50.0,
    "70米风速": 70.0,
    "轮毂高度风速": 80.0,
    "10米风向": 10.0,
    "30米风向": 30.0,
    "50米风向": 50.0,
    "70米风向": 70.0,
    "轮毂高度风向": 80.0,
}

TIME_FEATURE_COLS = (
    "minute_sin",
    "minute_cos",
    "dow_sin",
    "dow_cos",
    "doy_sin",
    "doy_cos",
    "month_sin",
    "month_cos",
)

FEATURE_SCHEMA = (
    "10米风速",
    "30米风速",
    "50米风速",
    "70米风速",
    "轮毂高度风速",
    "10m气温",
    "10m气压",
    "10m湿度",
    "10米风向_sin",
    "10米风向_cos",
    "30米风向_sin",
    "30米风向_cos",
    "50米风向_sin",
    "50米风向_cos",
    "70米风向_sin",
    "70米风向_cos",
    "轮毂高度风向_sin",
    "轮毂高度风向_cos",
    "minute_sin",
    "minute_cos",
    "dow_sin",
    "dow_cos",
    "doy_sin",
    "doy_cos",
    "month_sin",
    "month_cos",
    "10米风速_sq",
    "10米风速_cube",
    "30米风速_sq",
    "30米风速_cube",
    "50米风速_sq",
    "50米风速_cube",
    "70米风速_sq",
    "70米风速_cube",
    "轮毂高度风速_sq",
    "轮毂高度风速_cube",
    "轮毂高度风速_minus_10米风速",
    "轮毂高度风速_ratio_10米风速",
    "轮毂高度风速_minus_30米风速",
    "轮毂高度风速_ratio_30米风速",
    "轮毂高度风速_minus_50米风速",
    "轮毂高度风速_ratio_50米风速",
    "轮毂高度风速_minus_70米风速",
    "轮毂高度风速_ratio_70米风速",
    "功率",
)

FEATURE_GROUPS = {
    "P": (
        "power_last",
        "power_mean_4",
        "power_mean_16",
        "power_mean_32",
        "power_slope_4",
        "power_std_4",
        "power_mean_abs_step_4",
        "power_slope_8",
        "power_std_8",
        "power_mean_abs_step_8",
        "power_slope_16",
        "power_std_16",
        "power_mean_abs_step_16",
        "power_slope_32",
        "power_std_32",
        "power_mean_abs_step_32",
        "power_range_16",
        "power_range_32",
        "power_low_fraction_16",
        "power_low_fraction_32",
    ),
    "H": (
        "hub_wind_last",
        "hub_wind_mean_4",
        "hub_wind_mean_16",
        "hub_wind_slope_4",
        "hub_wind_std_4",
        "hub_wind_slope_8",
        "hub_wind_std_8",
        "hub_wind_slope_16",
        "hub_wind_std_16",
        "hub_wind_slope_32",
        "hub_wind_std_32",
        "hub_wind_mean_abs_step_16",
    ),
    "M": (
        "all_height_wind_last_mean",
        "all_height_wind_last_std",
        "hub_minus_height_mean",
    ),
    "D": (
        "direction_turn_lag_1",
        "direction_turn_lag_4",
        "direction_turn_lag_16",
        "direction_mean_turn_16",
    ),
    "C": (
        "power_wind_slope_product_8",
        "power_wind_slope_product_16",
        "power_minus_wind_cube_proxy",
        "power_wind_change_correlation_16",
    ),
}
FULL_REGIME_FEATURE_NAMES = tuple(
    name for group in ("P", "H", "M", "D", "C") for name in FEATURE_GROUPS[group]
)
F7_FEATURE_NAMES = tuple(
    name for group in ("P", "H", "D") for name in FEATURE_GROUPS[group]
)


@dataclass(frozen=True)
class TimestampRule:
    semantics: str
    sampling_minutes: int
    publication_delay_minutes: int = 0
    evidence: str = "conservative_default_no_authoritative_metadata"

    @property
    def availability_shift_minutes(self) -> int:
        if self.semantics in {"interval_start", "assumed_interval_start"}:
            return self.sampling_minutes + self.publication_delay_minutes
        if self.semantics in {"interval_end", "instantaneous"}:
            return self.publication_delay_minutes
        raise ValueError(f"不支持的timestamp semantics: {self.semantics}")


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_file(path: os.PathLike[str] | str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def sha256_json(value: Any) -> str:
    payload = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def atomic_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with open(tmp, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2, default=str)
    os.replace(tmp, path)


def assert_raw_path(path: Path) -> None:
    resolved = path.resolve()
    forbidden = (RAW_ROOT / "processed_npz").resolve()
    if forbidden == resolved or forbidden in resolved.parents:
        raise ValueError(f"正式Round-3流水线禁止读取processed_npz: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"原始输入必须是xlsx: {path}")


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("（", "(").replace("）", ")").replace("°", "")
    return re.sub(r"[\s_\-—/\\()]+", "", text)


def canonical_header(value: Any, kind: str) -> str | None:
    text = normalize_header(value)
    if not text:
        return None
    if "时间" in text or text in {"time", "datetime", "timestamp"}:
        return "时间"
    if kind == "power":
        if "功率" in text or "出力" in text:
            return TARGET_COL
        return None
    height_aliases = (
        ("轮毂", "轮毂高度"),
        ("70米", "70米"),
        ("50米", "50米"),
        ("30米", "30米"),
        ("10米", "10米"),
    )
    for token, prefix in height_aliases:
        if token in text and "风速" in text:
            return f"{prefix}风速"
        if token in text and "风向" in text:
            return f"{prefix}风向"
    if "气温" in text or "温度" in text:
        return "10m气温"
    if "气压" in text:
        return "10m气压"
    if "湿度" in text:
        return "10m湿度"
    return None


def parse_datetime(value: Any) -> pd.Timestamp | None:
    if value is None:
        return None
    try:
        timestamp = pd.Timestamp(value)
    except (TypeError, ValueError, OverflowError):
        return None
    if pd.isna(timestamp):
        return None
    if timestamp.tzinfo is not None:
        timestamp = timestamp.tz_convert(None)
    if timestamp.year < 2000 or timestamp.year > 2100:
        return None
    return timestamp


def numeric(value: Any) -> float:
    if value is None:
        return math.nan
    if isinstance(value, str):
        stripped = value.strip().lower()
        if stripped in {"", "null", "none", "nan", "n/a", "na", "--"}:
            return math.nan
    try:
        result = float(value)
    except (TypeError, ValueError):
        return math.nan
    return result if np.isfinite(result) else math.nan


def read_xlsx_table(path: Path, kind: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Read heterogeneous multi-sheet workbooks without scanning formatted tails."""
    assert_raw_path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    sheet_audit: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            header_map: dict[int, str] | None = None
            header_row_number: int | None = None
            rows_seen = 0
            valid_rows = 0
            blank_streak = 0
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                rows_seen += 1
                if header_map is None:
                    candidate: dict[int, str] = {}
                    for index, value in enumerate(row):
                        canonical = canonical_header(value, kind)
                        if canonical is not None and canonical not in candidate.values():
                            candidate[index] = canonical
                    if "时间" in candidate.values() and (
                        TARGET_COL in candidate.values()
                        if kind == "power"
                        else any(name in candidate.values() for name in SPEED_COLS)
                    ):
                        header_map = candidate
                        header_row_number = row_number
                    elif row_number >= 50:
                        break
                    continue

                selected = {
                    name: (row[index] if index < len(row) else None)
                    for index, name in header_map.items()
                }
                timestamp = parse_datetime(selected.get("时间"))
                if timestamp is None:
                    if all(
                        selected.get(name) is None
                        for name in header_map.values()
                        if name != "时间"
                    ):
                        blank_streak += 1
                        if valid_rows and blank_streak >= BLANK_ROW_STOP:
                            break
                    continue
                blank_streak = 0
                item: dict[str, Any] = {"时间": timestamp}
                for name in header_map.values():
                    if name != "时间":
                        item[name] = numeric(selected.get(name))
                records.append(item)
                valid_rows += 1
            sheet_audit.append(
                {
                    "sheet": worksheet.title,
                    "header_row": header_row_number,
                    "rows_scanned": rows_seen,
                    "valid_rows": valid_rows,
                    "declared_max_row": worksheet.max_row,
                    "declared_max_column": worksheet.max_column,
                }
            )
    finally:
        workbook.close()
    if not records:
        raise ValueError(f"未从{path}读取到{kind}有效数据")
    frame = pd.DataFrame.from_records(records)
    frame["时间"] = pd.to_datetime(frame["时间"])
    audit = {
        "path": str(path.resolve()),
        "sha256": sha256_file(path),
        "kind": kind,
        "records": len(frame),
        "sheets": sheet_audit,
        "columns": list(frame.columns),
    }
    return frame, audit


def infer_interval_minutes(index: pd.DatetimeIndex, fallback: int) -> int:
    unique = pd.DatetimeIndex(index.drop_duplicates()).sort_values()
    if len(unique) < 2:
        return fallback
    diffs = np.diff(unique.view("i8")) / 60_000_000_000
    diffs = diffs[(diffs > 0) & (diffs <= 60)]
    if not len(diffs):
        return fallback
    rounded = np.rint(diffs).astype(int)
    values, counts = np.unique(rounded, return_counts=True)
    candidate = int(values[np.argmax(counts)])
    return candidate if candidate in {1, 5, 10, 15, 30, 60} else fallback


def load_semantics_config(path: str | None) -> dict[str, Any]:
    if not path:
        return {}
    config_path = Path(path)
    forbidden = (RAW_ROOT / "processed_npz").resolve()
    resolved = config_path.resolve()
    if resolved == forbidden or forbidden in resolved.parents:
        raise ValueError("timestamp semantics配置也禁止从processed_npz读取")
    with open(config_path, "r", encoding="utf-8") as handle:
        config = json.load(handle)
    if not isinstance(config, dict):
        raise ValueError("timestamp semantics配置必须是JSON对象")
    return config


def timestamp_rule(
    farm_id: str,
    stream: str,
    sampling_minutes: int,
    config: dict[str, Any],
) -> TimestampRule:
    station = config.get(farm_id, {}) if isinstance(config.get(farm_id, {}), dict) else {}
    source = station.get(stream, {}) if isinstance(station.get(stream, {}), dict) else {}
    semantics = source.get("semantics", "assumed_interval_start")
    delay = int(source.get("publication_delay_minutes", 0))
    evidence = source.get("evidence", "conservative_default_no_authoritative_metadata")
    return TimestampRule(
        semantics=semantics,
        sampling_minutes=int(source.get("sampling_minutes", sampling_minutes)),
        publication_delay_minutes=delay,
        evidence=str(evidence),
    )


def apply_available_time(
    frame: pd.DataFrame,
    rule: TimestampRule,
) -> pd.DataFrame:
    result = frame.copy()
    result["原始时间"] = pd.to_datetime(result["时间"])
    result["时间"] = result["原始时间"] + pd.to_timedelta(
        rule.availability_shift_minutes, unit="min"
    )
    result = result.sort_values("时间")
    return result


def circular_mean_degrees(values: Iterable[float]) -> float:
    array = np.asarray(list(values), dtype=float)
    array = array[np.isfinite(array)]
    if not len(array):
        return math.nan
    radians = np.deg2rad(np.mod(array, 360.0))
    vector = np.mean(np.exp(1j * radians))
    if abs(vector) < 1e-8:
        return math.nan
    return float(np.mod(np.rad2deg(np.angle(vector)), 360.0))


def deduplicate_power(frame: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    source = frame[["时间", TARGET_COL]].copy()
    duplicate_rows = int(source.duplicated("时间", keep=False).sum())
    conflict_count = 0
    if duplicate_rows:
        grouped = source.groupby("时间", sort=True)[TARGET_COL]
        conflict_count = int(
            grouped.apply(
                lambda values: np.nanmax(values) - np.nanmin(values)
                if np.isfinite(values).sum() >= 2
                else 0.0
            ).gt(1e-9).sum()
        )
    result = source.groupby("时间", as_index=True)[TARGET_COL].median().to_frame()
    return result, {
        "duplicate_rows": duplicate_rows,
        "duplicate_timestamps": int(source["时间"].duplicated().sum()),
        "conflicting_duplicate_timestamps": conflict_count,
    }


def deduplicate_weather(frame: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    columns = [name for name in (*SPEED_COLS, *DIRECTION_BASE_COLS, *MET_COLS) if name in frame]
    work = frame[["时间", *columns]].copy()
    duplicate_rows = int(work.duplicated("时间", keep=False).sum())
    scalar_cols = [name for name in (*SPEED_COLS, *MET_COLS) if name in work]
    grouped_scalar = work.groupby("时间", sort=True)[scalar_cols].mean()
    result = grouped_scalar.copy()
    for name in DIRECTION_BASE_COLS:
        if name in work:
            result[name] = work.groupby("时间", sort=True)[name].apply(
                circular_mean_degrees
            )
    return result.sort_index(), {
        "duplicate_rows": duplicate_rows,
        "duplicate_timestamps": int(work["时间"].duplicated().sum()),
    }


def clean_power(series: pd.Series) -> tuple[pd.Series, dict[str, int]]:
    values = pd.to_numeric(series, errors="coerce").astype(float)
    sentinel = values <= -90.0
    nonfinite = ~np.isfinite(values)
    small_negative = (values < 0.0) & ~sentinel
    values = values.mask(sentinel | nonfinite)
    values = values.clip(lower=0.0)
    return values, {
        "power_sentinel_count": int(sentinel.sum()),
        "power_nonfinite_count": int(nonfinite.sum()),
        "power_negative_clipped_count": int(small_negative.sum()),
    }


def clean_weather(frame: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, Any]]:
    result = frame.copy()
    audit: dict[str, Any] = {}
    for column in result.columns:
        result[column] = pd.to_numeric(result[column], errors="coerce").astype(float)
        sentinel = np.isclose(result[column], -99.0, equal_nan=False) | np.isclose(
            result[column], 3276.7, atol=0.05, equal_nan=False
        )
        audit[f"{column}_sentinel_count"] = int(sentinel.sum())
        result[column] = result[column].mask(sentinel)
    for column in SPEED_COLS:
        if column in result:
            invalid = ~result[column].between(0.0, 60.0)
            audit[f"{column}_physical_invalid_count"] = int(invalid.sum())
            result[column] = result[column].mask(invalid)
    for column in DIRECTION_BASE_COLS:
        if column in result:
            invalid = ~result[column].between(0.0, 360.0)
            audit[f"{column}_physical_invalid_count"] = int(invalid.sum())
            result[column] = result[column].mask(invalid)
            result[column] = np.mod(result[column], 360.0)
    if "10m气温" in result:
        invalid = ~result["10m气温"].between(-50.0, 60.0)
        audit["temperature_invalid_count"] = int(invalid.sum())
        result["10m气温"] = result["10m气温"].mask(invalid)
    if "10m气压" in result:
        scaled = result["10m气压"].between(2000.0, 20000.0)
        result.loc[scaled, "10m气压"] = result.loc[scaled, "10m气压"] / 10.0
        invalid = ~result["10m气压"].between(800.0, 1150.0)
        audit["pressure_divide10_count"] = int(scaled.sum())
        audit["pressure_invalid_count"] = int(invalid.sum())
        result["10m气压"] = result["10m气压"].mask(invalid)
    if "10m湿度" in result:
        invalid = ~result["10m湿度"].between(0.0, 100.0)
        audit["humidity_invalid_count"] = int(invalid.sum())
        result["10m湿度"] = result["10m湿度"].mask(invalid)
    return result, audit


def aggregate_power_15min(frame: pd.DataFrame) -> pd.DataFrame:
    """Causally aggregate power on the availability-time axis.

    Power is normally already 15-minute data.  Resampling is still explicit so
    that duplicate/off-grid records cannot silently disappear during reindex.
    A label at ``t`` only contains records whose availability time is in
    ``(t-15min, t]``.
    """
    return (
        frame[[TARGET_COL]]
        .resample(TIME_FREQ, closed="right", label="right")
        .median()
        .sort_index()
    )


def aggregate_weather_15min(frame: pd.DataFrame) -> pd.DataFrame:
    scalar_cols = [name for name in (*SPEED_COLS, *MET_COLS) if name in frame]
    output = frame[scalar_cols].resample(
        TIME_FREQ, closed="right", label="right"
    ).mean()
    for name in DIRECTION_BASE_COLS:
        if name not in frame:
            continue
        radians = np.deg2rad(frame[name])
        sin_mean = pd.Series(np.sin(radians), index=frame.index).resample(
            TIME_FREQ, closed="right", label="right"
        ).mean()
        cos_mean = pd.Series(np.cos(radians), index=frame.index).resample(
            TIME_FREQ, closed="right", label="right"
        ).mean()
        magnitude = np.sqrt(sin_mean**2 + cos_mean**2)
        angle = np.mod(np.rad2deg(np.arctan2(sin_mean, cos_mean)), 360.0)
        output[name] = angle.where(magnitude >= 1e-6)
    return output.sort_index()


def reliability_scalar(series: pd.Series, train_mask: np.ndarray) -> dict[str, Any]:
    train = pd.to_numeric(series.iloc[np.flatnonzero(train_mask)], errors="coerce")
    valid = train.dropna()
    fraction = float(len(valid) / max(1, len(train)))
    std = float(valid.std()) if len(valid) >= 2 else 0.0
    unique = int(valid.round(5).nunique())
    reliable = bool(fraction >= 0.05 and std > 1e-4 and unique >= 4)
    return {
        "valid_fraction_train": fraction,
        "std_train": std,
        "unique_train": unique,
        "reliable_train": reliable,
    }


def reliability_direction(series: pd.Series, train_mask: np.ndarray) -> dict[str, Any]:
    train = pd.to_numeric(series.iloc[np.flatnonzero(train_mask)], errors="coerce")
    valid = train.dropna()
    fraction = float(len(valid) / max(1, len(train)))
    unique = int(valid.round(2).nunique())
    if len(valid):
        vector = np.mean(np.exp(1j * np.deg2rad(valid.to_numpy())))
        circular_variance = float(1.0 - abs(vector))
    else:
        circular_variance = 0.0
    reliable = bool(fraction >= 0.05 and unique >= 4 and circular_variance > 1e-5)
    return {
        "valid_fraction_train": fraction,
        "unique_train": unique,
        "circular_variance_train": circular_variance,
        "reliable_train": reliable,
    }


def nearest_columns(target: str, candidates: list[str]) -> list[str]:
    height = HEIGHTS[target]
    return sorted(candidates, key=lambda name: (abs(HEIGHTS[name] - height), HEIGHTS[name]))


def reconstruct_speeds(
    frame: pd.DataFrame,
    train_mask: np.ndarray,
) -> tuple[pd.DataFrame, np.ndarray, dict[str, Any]]:
    result = pd.DataFrame(index=frame.index)
    reliability: dict[str, Any] = {}
    reliable: list[str] = []
    for name in SPEED_COLS:
        source = frame[name] if name in frame else pd.Series(np.nan, index=frame.index)
        info = reliability_scalar(source, train_mask)
        reliability[name] = info
        if info["reliable_train"]:
            reliable.append(name)
        result[name] = source
    if not reliable:
        raise ValueError("训练段没有任何可靠风速传感器")

    methods: dict[str, str] = {}
    for target in SPEED_COLS:
        if target in reliable:
            methods[target] = "own_sensor_then_same_time_linear_height_interpolation"
        else:
            methods[target] = "structural_same_time_linear_height_interpolation"
            # Values from a train-diagnosed failed sensor must never leak back
            # through the reconstruction path.
            result[target] = np.nan
        missing = result[target].isna()
        candidates = [name for name in reliable if name != target]
        if missing.any() and candidates:
            target_height = HEIGHTS[target]
            lower_value = np.full(len(result), np.nan, dtype=float)
            lower_height = np.full(len(result), np.nan, dtype=float)
            upper_value = np.full(len(result), np.nan, dtype=float)
            upper_height = np.full(len(result), np.nan, dtype=float)
            lower_names = sorted(
                (name for name in candidates if HEIGHTS[name] <= target_height),
                key=lambda name: HEIGHTS[name],
                reverse=True,
            )
            upper_names = sorted(
                (name for name in candidates if HEIGHTS[name] >= target_height),
                key=lambda name: HEIGHTS[name],
            )
            for candidate in lower_names:
                values = frame[candidate].to_numpy(float)
                use = ~np.isfinite(lower_value) & np.isfinite(values)
                lower_value[use] = values[use]
                lower_height[use] = HEIGHTS[candidate]
            for candidate in upper_names:
                values = frame[candidate].to_numpy(float)
                use = ~np.isfinite(upper_value) & np.isfinite(values)
                upper_value[use] = values[use]
                upper_height[use] = HEIGHTS[candidate]

            proxy = np.full(len(result), np.nan, dtype=float)
            both = np.isfinite(lower_value) & np.isfinite(upper_value)
            distinct = both & (upper_height > lower_height)
            ratio = np.zeros(len(result), dtype=float)
            ratio[distinct] = (
                (target_height - lower_height[distinct])
                / (upper_height[distinct] - lower_height[distinct])
            )
            proxy[distinct] = lower_value[distinct] + ratio[distinct] * (
                upper_value[distinct] - lower_value[distinct]
            )
            same = both & ~distinct
            proxy[same] = lower_value[same]
            lower_only = np.isfinite(lower_value) & ~np.isfinite(proxy)
            upper_only = np.isfinite(upper_value) & ~np.isfinite(proxy)
            proxy[lower_only] = lower_value[lower_only]
            proxy[upper_only] = upper_value[upper_only]

            # If all reliable sensors happen to lie on only one side, the
            # nearest same-time height is a conservative non-extrapolating
            # fallback.
            if not np.isfinite(proxy).all():
                for candidate in nearest_columns(target, candidates):
                    values = frame[candidate].to_numpy(float)
                    use = ~np.isfinite(proxy) & np.isfinite(values)
                    proxy[use] = values[use]
            use = missing.to_numpy() & np.isfinite(proxy)
            result.loc[use, target] = proxy[use]
    result = result.ffill(limit=CAUSAL_FILL_LIMIT)
    essential_valid = result[list(SPEED_COLS)].notna().any(axis=1).to_numpy(bool)
    train_positions = np.flatnonzero(train_mask)
    medians: dict[str, float] = {}
    for name in SPEED_COLS:
        median = float(result[name].iloc[train_positions].median())
        if not np.isfinite(median):
            median = float(
                np.nanmedian(result[list(reliable)].iloc[train_positions].to_numpy())
            )
        if not np.isfinite(median):
            raise ValueError(f"无法为{name}建立训练段回退值")
        medians[name] = median
        result[name] = result[name].fillna(median)
    return result, essential_valid, {
        "reliability": reliability,
        "reconstruction_method": methods,
        "train_medians": medians,
        "reliable_columns": reliable,
    }


def reconstruct_directions(
    frame: pd.DataFrame,
    train_mask: np.ndarray,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    raw = pd.DataFrame(index=frame.index)
    reliability: dict[str, Any] = {}
    reliable: list[str] = []
    for name in DIRECTION_BASE_COLS:
        source = frame[name] if name in frame else pd.Series(np.nan, index=frame.index)
        info = reliability_direction(source, train_mask)
        reliability[name] = info
        if info["reliable_train"]:
            reliable.append(name)
        raw[name] = source

    output = raw.copy()
    methods: dict[str, str] = {}
    for target in DIRECTION_BASE_COLS:
        methods[target] = (
            "own_sensor_then_weighted_circular_height_proxy"
            if target in reliable
            else "structural_weighted_circular_height_proxy"
        )
        values = output[target].to_numpy(float)
        if target not in reliable:
            values[:] = np.nan
        missing = ~np.isfinite(values)
        candidates = [name for name in reliable if name != target]
        if candidates and missing.any():
            vectors = np.zeros(len(values), dtype=np.complex128)
            weights = np.zeros(len(values), dtype=float)
            for candidate in candidates:
                angles = raw[candidate].to_numpy(float)
                valid = np.isfinite(angles) & missing
                weight = 1.0 / max(1.0, abs(HEIGHTS[candidate] - HEIGHTS[target]))
                vectors[valid] += weight * np.exp(1j * np.deg2rad(angles[valid]))
                weights[valid] += weight
            valid_proxy = missing & (weights > 0) & (np.abs(vectors) > 1e-8)
            values[valid_proxy] = np.mod(
                np.rad2deg(np.angle(vectors[valid_proxy])), 360.0
            )
        output[target] = values

    output = output.ffill(limit=CAUSAL_FILL_LIMIT)
    fallback: dict[str, float | None] = {}
    structurally_missing: list[str] = []
    unknown_masks: dict[str, np.ndarray] = {}
    for name in DIRECTION_BASE_COLS:
        # A long missing direction is "unknown", not the circular mean of the
        # training period.  Encoding it as the neutral vector (0, 0) avoids
        # inventing a preferred direction and keeps validation/test transforms
        # independent of their own distributions.
        fallback[name] = None
        unknown_masks[name] = output[name].isna().to_numpy()
        if not reliability[name]["reliable_train"]:
            structurally_missing.append(name)

    encoded = pd.DataFrame(index=output.index)
    violations_before = 0
    violations_after = 0
    unknown_count = 0
    total_pairs = len(output) * len(DIRECTION_BASE_COLS)
    for name in DIRECTION_BASE_COLS:
        unknown = unknown_masks[name]
        angle = output[name].to_numpy(float)
        radians = np.deg2rad(angle)
        sin_value = np.sin(radians)
        cos_value = np.cos(radians)
        # A failed *source* sensor may still be reconstructed from a reliable
        # neighbouring height.  Only rows with no causal/cross-height estimate
        # receive the neutral unknown-direction vector.
        sin_value[unknown] = 0.0
        cos_value[unknown] = 0.0
        norm_error = np.abs(sin_value**2 + cos_value**2 - 1.0)
        physical = ~unknown
        violations_before += int(np.sum(physical & (norm_error > 1e-6)))
        magnitude = np.sqrt(sin_value**2 + cos_value**2)
        repair = physical & (magnitude > 1e-12)
        sin_value[repair] /= magnitude[repair]
        cos_value[repair] /= magnitude[repair]
        after = np.abs(sin_value**2 + cos_value**2 - 1.0)
        violations_after += int(np.sum(physical & (after > 1e-6)))
        unknown_count += int(unknown.sum())
        encoded[f"{name}_sin"] = sin_value
        encoded[f"{name}_cos"] = cos_value
    if violations_after:
        raise ValueError(f"风向单位圆修复后仍有{violations_after}个违规点")
    return encoded, {
        "reliability": reliability,
        "reliable_columns": reliable,
        "reconstruction_method": methods,
        "train_circular_fallback": fallback,
        "structurally_missing": structurally_missing,
        "unknown_rate_per_mille": 1000.0 * unknown_count / max(1, total_pairs),
        "unity_violations_before_repair_per_mille": (
            1000.0 * violations_before / max(1, total_pairs)
        ),
        "unity_violations_after_repair_per_mille": (
            1000.0 * violations_after / max(1, total_pairs)
        ),
    }


def fill_meteorology(
    frame: pd.DataFrame,
    train_mask: np.ndarray,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    defaults = {"10m气温": 15.0, "10m气压": 1013.25, "10m湿度": 50.0}
    result = pd.DataFrame(index=frame.index)
    medians: dict[str, float] = {}
    reliability: dict[str, Any] = {}
    train_positions = np.flatnonzero(train_mask)
    for name in MET_COLS:
        source = frame[name] if name in frame else pd.Series(np.nan, index=frame.index)
        info = reliability_scalar(source, train_mask)
        reliability[name] = info
        if not info["reliable_train"]:
            source = pd.Series(np.nan, index=frame.index, dtype=float)
        source = source.ffill(limit=CAUSAL_FILL_LIMIT)
        median = float(source.iloc[train_positions].median())
        if not np.isfinite(median):
            median = defaults[name]
        medians[name] = median
        result[name] = source.fillna(median)
    return result, {
        "train_medians": medians,
        "reliability": reliability,
        "defaults": defaults,
    }


def add_time_features(frame: pd.DataFrame) -> None:
    index = frame.index
    minute = index.hour * 60 + index.minute
    frame["minute_sin"] = np.sin(2.0 * np.pi * minute / 1440.0)
    frame["minute_cos"] = np.cos(2.0 * np.pi * minute / 1440.0)
    frame["dow_sin"] = np.sin(2.0 * np.pi * index.dayofweek / 7.0)
    frame["dow_cos"] = np.cos(2.0 * np.pi * index.dayofweek / 7.0)
    frame["doy_sin"] = np.sin(2.0 * np.pi * index.dayofyear / 366.0)
    frame["doy_cos"] = np.cos(2.0 * np.pi * index.dayofyear / 366.0)
    frame["month_sin"] = np.sin(2.0 * np.pi * index.month / 12.0)
    frame["month_cos"] = np.cos(2.0 * np.pi * index.month / 12.0)


def add_wind_features(frame: pd.DataFrame) -> None:
    for name in SPEED_COLS:
        frame[f"{name}_sq"] = frame[name] ** 2
        frame[f"{name}_cube"] = frame[name] ** 3
    hub = "轮毂高度风速"
    for name in SPEED_COLS[:-1]:
        frame[f"{hub}_minus_{name}"] = frame[hub] - frame[name]
        frame[f"{hub}_ratio_{name}"] = frame[hub] / frame[name].clip(lower=0.5)


def source_feature_availability(
    speed_audit: dict[str, Any],
    direction_audit: dict[str, Any],
    meteorology_audit: dict[str, Any],
) -> tuple[int, list[str]]:
    """Describe which fixed-schema channels lack a direct reliable source.

    These channels are still reconstructed causally before model input.  The
    audit distinguishes that reconstruction from an originally observed
    45-channel station and fulfils the fixed-schema transparency requirement.
    """
    missing: set[str] = set()
    hub = "轮毂高度风速"
    for name, info in speed_audit["reliability"].items():
        if info["reliable_train"]:
            continue
        missing.update((name, f"{name}_sq", f"{name}_cube"))
        if name == hub:
            for lower in SPEED_COLS[:-1]:
                missing.add(f"{hub}_minus_{lower}")
                missing.add(f"{hub}_ratio_{lower}")
        else:
            missing.add(f"{hub}_minus_{name}")
            missing.add(f"{hub}_ratio_{name}")
    for name, info in direction_audit["reliability"].items():
        if not info["reliable_train"]:
            missing.update((f"{name}_sin", f"{name}_cos"))
    for name, info in meteorology_audit["reliability"].items():
        if not info["reliable_train"]:
            missing.add(name)
    ordered = [name for name in FEATURE_SCHEMA if name in missing]
    return len(FEATURE_SCHEMA) - len(ordered), ordered


def make_origins(
    power_valid: np.ndarray,
    essential_wind_valid: np.ndarray,
    split_code: np.ndarray,
) -> dict[str, np.ndarray]:
    output: dict[str, list[int]] = {"train": [], "val": [], "test": []}
    names = {0: "train", 1: "val", 2: "test"}
    for origin in range(HISTORY_LEN, len(power_valid) - FORECAST_LEN + 1):
        history = slice(origin - HISTORY_LEN, origin)
        target = slice(origin, origin + FORECAST_LEN)
        if not power_valid[history].all() or not power_valid[target].all():
            continue
        if not essential_wind_valid[history].all():
            continue
        codes = split_code[target]
        if not np.all(codes == codes[0]):
            continue
        output[names[int(codes[0])]].append(origin)
    return {name: np.asarray(values, dtype=np.int64) for name, values in output.items()}


def feasibility_label(
    eligible_windows: int,
    calendar_days: float,
    retention_ratio: float,
) -> str:
    # The primary tier is deliberately determined only by eligible training
    # windows, matching the frozen Round-3 protocol.  Calendar coverage and
    # retention remain separate audit fields rather than silently changing a
    # station's tier.
    del calendar_days, retention_ratio
    if eligible_windows >= 50_000:
        return "sufficient"
    if eligible_windows >= 20_000:
        return "limited"
    if eligible_windows >= 5_000:
        return "constrained"
    return "insufficient"


def build_regime_config(
    scaler_x: StandardScaler,
    power_reference: float,
) -> dict[str, Any]:
    columns = list(FEATURE_SCHEMA)
    speed_indices = [columns.index(name) for name in SPEED_COLS]
    target_index = columns.index(TARGET_COL)
    sin_index = columns.index("轮毂高度风向_sin")
    cos_index = columns.index("轮毂高度风向_cos")
    config = {
        "target_channel_index": target_index,
        "power_mean": float(scaler_x.mean_[target_index]),
        "power_scale": float(scaler_x.scale_[target_index]),
        # Compatibility name expected by the existing F7 layer.  This value is
        # explicitly a train-only q99.9 reference, not installed capacity.
        "capacity": float(power_reference),
        "power_reference_kind": "train_power_q999",
        "wind_speed_indices": speed_indices,
        "wind_speed_names": list(SPEED_COLS),
        "wind_speed_means": [float(scaler_x.mean_[i]) for i in speed_indices],
        "wind_speed_scales": [float(scaler_x.scale_[i]) for i in speed_indices],
        "hub_wind_position": 4,
        "direction_sin_index": sin_index,
        "direction_cos_index": cos_index,
        "direction_sin_mean": float(scaler_x.mean_[sin_index]),
        "direction_sin_scale": float(scaler_x.scale_[sin_index]),
        "direction_cos_mean": float(scaler_x.mean_[cos_index]),
        "direction_cos_scale": float(scaler_x.scale_[cos_index]),
        "windows": [4, 8, 16, 32],
        "low_power_threshold": 0.02,
        "wind_speed_normalizer": 25.0,
        "feature_names": list(FULL_REGIME_FEATURE_NAMES),
        "selected_groups": ["P", "H", "D"],
        "selected_feature_names": list(F7_FEATURE_NAMES),
        "missing_safe_direction_required": True,
    }
    if len(F7_FEATURE_NAMES) != 36:
        raise AssertionError("F7特征数必须为36")
    return config


def result_paths(farm_id: str) -> dict[str, Path]:
    return {
        "array": RESULT_ROOT / "prepared_data" / "feature_arrays" / f"{farm_id}.npz",
        "canonical": (
            RESULT_ROOT / "prepared_data" / "canonical_15min" / f"{farm_id}.parquet"
        ),
        "split_indices": (
            RESULT_ROOT / "prepared_data" / "split_indices" / f"{farm_id}.npz"
        ),
        "bundle": RESULT_ROOT / "preprocess" / farm_id / "preprocessing_bundle.joblib",
        "power_reference": (
            RESULT_ROOT / "preprocess" / farm_id / "power_reference.json"
        ),
        "regime_config": (
            RESULT_ROOT / "preprocess" / farm_id / "regime_feature_config.json"
        ),
        "farm_manifest": RESULT_ROOT / "manifests" / "preprocess" / f"{farm_id}.json",
    }


def prepare_farm(
    farm_id: str,
    semantics_config: dict[str, Any],
    force: bool = False,
) -> dict[str, Any]:
    if farm_id not in EXPECTED_FARMS:
        raise ValueError(f"未知场站: {farm_id}")
    paths = result_paths(farm_id)
    if paths["farm_manifest"].exists() and not force:
        with open(paths["farm_manifest"], "r", encoding="utf-8") as handle:
            prior = json.load(handle)
        summary = prior.get("summary", {})
        array_ok = (
            paths["array"].is_file()
            and summary.get("array_sha256") == sha256_file(paths["array"])
        )
        bundle_ok = (
            paths["bundle"].is_file()
            and summary.get("bundle_sha256") == sha256_file(paths["bundle"])
        )
        split_ok = (
            paths["split_indices"].is_file()
            and summary.get("split_indices_sha256")
            == sha256_file(paths["split_indices"])
        )
        json_ok = paths["power_reference"].is_file() and paths[
            "regime_config"
        ].is_file()
        if (
            prior.get("status") == "complete"
            and array_ok
            and bundle_ok
            and split_ok
            and json_ok
        ):
            print(f"[resume] {farm_id} 已完成，跳过")
            return summary
        print(f"[stale] {farm_id}已有manifest但产物不完整或hash不符，重新生成")

    farm_root = RAW_ROOT / farm_id
    power_candidates = sorted(farm_root.glob("*场站出力*.xlsx"))
    weather_candidates = sorted(
        list(farm_root.glob("*测风数据*.xlsx"))
        + list(farm_root.glob("*测风塔数据*.xlsx"))
    )
    run_record_candidates = sorted(farm_root.glob("*运行记录*.xlsx"))
    if len(power_candidates) != 1 or len(weather_candidates) != 1:
        raise FileNotFoundError(
            f"{farm_id}原始文件不唯一: power={power_candidates}, weather={weather_candidates}"
        )
    run_record_audit: list[dict[str, Any]] = []
    for path in run_record_candidates:
        assert_raw_path(path)
        run_record_audit.append(
            {
                "path": str(path.resolve()),
                "sha256": sha256_file(path),
                "used_for_features": False,
                "used_as_nameplate_capacity": False,
            }
        )

    started = time.monotonic()
    power_raw, power_file_audit = read_xlsx_table(power_candidates[0], "power")
    weather_raw, weather_file_audit = read_xlsx_table(weather_candidates[0], "weather")
    power_minutes = infer_interval_minutes(
        pd.DatetimeIndex(power_raw["时间"]), fallback=15
    )
    weather_minutes = infer_interval_minutes(
        pd.DatetimeIndex(weather_raw["时间"]), fallback=15
    )
    power_rule = timestamp_rule(
        farm_id, "power", power_minutes, semantics_config
    )
    weather_rule = timestamp_rule(
        farm_id, "weather", weather_minutes, semantics_config
    )
    power_available = apply_available_time(power_raw, power_rule)
    weather_available = apply_available_time(weather_raw, weather_rule)

    # Clean individual observations before duplicate aggregation.  Otherwise a
    # sentinel and a valid duplicate could be averaged/median-combined into an
    # apparently valid but fabricated measurement.
    power_available[TARGET_COL], power_clean_audit = clean_power(
        power_available[TARGET_COL]
    )
    weather_columns = [
        name
        for name in (*SPEED_COLS, *DIRECTION_BASE_COLS, *MET_COLS)
        if name in weather_available
    ]
    cleaned_weather, weather_clean_audit = clean_weather(
        weather_available[weather_columns]
    )
    weather_available.loc[:, weather_columns] = cleaned_weather

    power, duplicate_power = deduplicate_power(power_available)
    weather, duplicate_weather = deduplicate_weather(weather_available)
    power = aggregate_power_15min(power)
    weather = aggregate_weather_15min(weather)

    common_start = max(power.index.min(), weather.index.min()).ceil(TIME_FREQ)
    common_end = min(power.index.max(), weather.index.max()).floor(TIME_FREQ)
    if common_end <= common_start:
        raise ValueError(f"{farm_id}功率与测风没有共同有效时间范围")
    grid = pd.date_range(common_start, common_end, freq=TIME_FREQ, name="时间")
    power = power.reindex(grid)
    weather = weather.reindex(grid)

    n_rows = len(grid)
    train_stop = int(math.floor(n_rows * 0.70))
    val_stop = int(math.floor(n_rows * 0.85))
    if train_stop <= HISTORY_LEN + FORECAST_LEN or val_stop <= train_stop:
        raise ValueError(f"{farm_id}数据量不足以切分")
    split_code = np.full(n_rows, 2, dtype=np.int8)
    split_code[:train_stop] = 0
    split_code[train_stop:val_stop] = 1
    train_mask = split_code == 0

    speed_frame, essential_wind_valid, speed_audit = reconstruct_speeds(
        weather, train_mask
    )
    direction_frame, direction_audit = reconstruct_directions(weather, train_mask)
    met_frame, met_audit = fill_meteorology(weather, train_mask)
    input_channels_available, missing_channels = source_feature_availability(
        speed_audit,
        direction_audit,
        met_audit,
    )

    power_values = power[TARGET_COL].astype(float)
    power_valid = np.isfinite(power_values.to_numpy())
    train_power = power_values.iloc[:train_stop].dropna().to_numpy(float)
    if len(train_power) < 100:
        raise ValueError(f"{farm_id}训练段有效功率不足: {len(train_power)}")
    power_reference = float(
        np.quantile(train_power, POWER_REFERENCE_QUANTILE, method="linear")
    )
    power_q99 = float(np.quantile(train_power, 0.99, method="linear"))
    power_train_max = float(np.max(train_power))
    if not np.isfinite(power_reference) or power_reference <= 1e-6:
        raise ValueError(f"{farm_id}训练段Q99.9功率参考无效: {power_reference}")

    features = pd.concat([speed_frame, met_frame, direction_frame], axis=1)
    add_time_features(features)
    add_wind_features(features)
    train_power_median = float(np.median(train_power))
    features[TARGET_COL] = power_values.fillna(train_power_median)
    missing_schema = [name for name in FEATURE_SCHEMA if name not in features]
    extra_schema = [name for name in features if name not in FEATURE_SCHEMA]
    if missing_schema:
        raise ValueError(f"{farm_id}缺少规范特征: {missing_schema}")
    if extra_schema:
        features = features.drop(columns=extra_schema)
    features = features.loc[:, list(FEATURE_SCHEMA)].astype(np.float64)
    if tuple(features.columns) != FEATURE_SCHEMA:
        raise AssertionError("FEATURE_SCHEMA列顺序漂移")
    if not np.isfinite(features.to_numpy()).all():
        bad = features.columns[~np.isfinite(features.to_numpy()).all(axis=0)].tolist()
        raise ValueError(f"{farm_id}规范特征仍有非有限值: {bad}")

    # Scaler statistics are estimated only from training timestamps which can
    # actually occur in eligible histories.  Invalid power and long wind gaps
    # are not allowed to bias the power channel toward its fallback value.
    scaler_fit_mask = train_mask & power_valid & essential_wind_valid
    if int(scaler_fit_mask.sum()) < 100:
        raise ValueError(
            f"{farm_id}训练段可用于Scaler拟合的时刻不足: "
            f"{int(scaler_fit_mask.sum())}"
        )
    scaler_x = StandardScaler()
    scaler_y = StandardScaler()
    scaler_x.fit(features.to_numpy()[scaler_fit_mask])
    scaler_y.fit(
        power_values.to_numpy()[scaler_fit_mask].reshape(-1, 1)
    )
    features_scaled = scaler_x.transform(features.to_numpy()).astype(np.float32)
    target_scaled = np.zeros(n_rows, dtype=np.float32)
    target_scaled[power_valid] = scaler_y.transform(
        power_values.to_numpy()[power_valid].reshape(-1, 1)
    ).ravel().astype(np.float32)
    target_mw = power_values.to_numpy(dtype=np.float32, na_value=np.nan)

    origins = make_origins(power_valid, essential_wind_valid, split_code)
    if any(len(origins[name]) == 0 for name in ("train", "val", "test")):
        raise ValueError(
            f"{farm_id}存在空窗口分区: "
            + ", ".join(f"{name}={len(value)}" for name, value in origins.items())
        )
    target_index_sets = {
        name: np.unique(
            (
                values[:, None]
                + np.arange(FORECAST_LEN, dtype=np.int64)[None, :]
            ).ravel()
        )
        for name, values in origins.items()
    }
    cross_split_target_overlap = {
        "train_validation": int(
            np.intersect1d(
                target_index_sets["train"], target_index_sets["val"]
            ).size
        ),
        "train_test": int(
            np.intersect1d(
                target_index_sets["train"], target_index_sets["test"]
            ).size
        ),
        "validation_test": int(
            np.intersect1d(
                target_index_sets["val"], target_index_sets["test"]
            ).size
        ),
    }
    if any(cross_split_target_overlap.values()):
        raise AssertionError(
            f"{farm_id}分区之间存在目标标签重叠: "
            f"{cross_split_target_overlap}"
        )
    theoretical_train = max(0, train_stop - HISTORY_LEN - FORECAST_LEN + 1)
    train_days = (
        (grid[train_stop - 1] - grid[0]).total_seconds() / 86400.0
        if train_stop > 1
        else 0.0
    )
    retention = len(origins["train"]) / max(1, theoretical_train)
    feasibility = feasibility_label(len(origins["train"]), train_days, retention)

    x_power_mean = float(scaler_x.mean_[-1])
    x_power_scale = float(scaler_x.scale_[-1])
    y_mean = float(scaler_y.mean_[0])
    y_scale = float(scaler_y.scale_[0])
    power_scale_ratio = x_power_scale / y_scale
    power_scale_offset = (x_power_mean - y_mean) / y_scale
    regime_config = build_regime_config(scaler_x, power_reference)
    schema_hash = sha256_json(list(FEATURE_SCHEMA))

    for path in paths.values():
        path.parent.mkdir(parents=True, exist_ok=True)
    power_reference_payload = {
        "protocol_version": PROTOCOL_VERSION,
        "farm_id": farm_id,
        "reference_kind": "train_power_q999",
        "reference_mw": power_reference,
        "quantile": POWER_REFERENCE_QUANTILE,
        "quantile_method": "linear",
        "train_valid_unique_power_count": int(len(train_power)),
        "train_power_q99_mw": power_q99,
        "train_power_q999_mw": power_reference,
        "train_power_max_mw": power_train_max,
        "source_split": "train",
        "imputed_values_used": False,
        "validation_or_test_used": False,
        "used_as_prediction_upper_clip": False,
    }
    power_reference_payload["reference_hash"] = sha256_json(
        power_reference_payload
    )
    atomic_json(paths["power_reference"], power_reference_payload)

    regime_payload = {
        "protocol_version": PROTOCOL_VERSION,
        "feature_schema_version": SCHEMA_VERSION,
        "schema_hash": schema_hash,
        "legacy_f7_schema_hash": LEGACY_F7_SCHEMA_HASH,
        "semantic_mapping_status": "exact_legacy_f7_schema_match",
        "farm_id": farm_id,
        "input_cols": list(FEATURE_SCHEMA),
        "target_index": 44,
        "scaler_x_mean": np.asarray(scaler_x.mean_, dtype=float).tolist(),
        "scaler_x_scale": np.asarray(scaler_x.scale_, dtype=float).tolist(),
        "scaler_y_mean": np.asarray(scaler_y.mean_, dtype=float).tolist(),
        "scaler_y_scale": np.asarray(scaler_y.scale_, dtype=float).tolist(),
        "power_scale_ratio": float(power_scale_ratio),
        "power_scale_offset": float(power_scale_offset),
        "power_reference_mw": power_reference,
        "power_reference_kind": "train_power_q999",
        "regime_feature_config": regime_config,
    }
    regime_payload["config_hash"] = sha256_json(regime_payload)
    atomic_json(paths["regime_config"], regime_payload)

    np.savez_compressed(
        paths["array"],
        features_scaled=features_scaled,
        target_scaled=target_scaled,
        target_mw=target_mw,
        timestamps_ns=grid.view("i8"),
        power_valid=power_valid.astype(np.uint8),
        essential_wind_valid=essential_wind_valid.astype(np.uint8),
        split_code=split_code,
        train_origins=origins["train"],
        val_origins=origins["val"],
        test_origins=origins["test"],
        input_cols=np.asarray(FEATURE_SCHEMA, dtype="U64"),
        schema_hash=np.asarray(schema_hash, dtype="U64"),
        target_index=np.asarray(44, dtype=np.int64),
        history_len=np.asarray(HISTORY_LEN, dtype=np.int64),
        forecast_len=np.asarray(FORECAST_LEN, dtype=np.int64),
        scaler_x_mean=np.asarray(scaler_x.mean_, dtype=np.float64),
        scaler_x_scale=np.asarray(scaler_x.scale_, dtype=np.float64),
        scaler_y_mean=np.asarray(scaler_y.mean_, dtype=np.float64),
        scaler_y_scale=np.asarray(scaler_y.scale_, dtype=np.float64),
        power_reference_mw=np.asarray([power_reference], dtype=np.float64),
        power_reference_kind=np.asarray("train_power_q999", dtype="U32"),
    )
    array_sha = sha256_file(paths["array"])
    np.savez_compressed(
        paths["split_indices"],
        train_origins=origins["train"],
        validation_origins=origins["val"],
        test_origins=origins["test"],
        history_len=np.asarray(HISTORY_LEN, dtype=np.int64),
        forecast_len=np.asarray(FORECAST_LEN, dtype=np.int64),
        train_stop_index=np.asarray(train_stop, dtype=np.int64),
        validation_stop_index=np.asarray(val_stop, dtype=np.int64),
    )

    canonical = features.copy()
    canonical[TARGET_COL] = target_mw
    canonical["power_valid"] = power_valid
    canonical["essential_wind_valid"] = essential_wind_valid
    canonical["split"] = np.asarray(["train", "validation", "test"])[split_code]
    canonical.to_parquet(paths["canonical"], compression="zstd", index=True)

    bundle = {
        "artifact_schema_version": ARTIFACT_SCHEMA_VERSION,
        "protocol_version": PROTOCOL_VERSION,
        "feature_schema_version": SCHEMA_VERSION,
        "schema_hash": schema_hash,
        "farm_id": farm_id,
        "input_cols": list(FEATURE_SCHEMA),
        "target_col": TARGET_COL,
        "target_index": 44,
        "history_len": HISTORY_LEN,
        "forecast_len": FORECAST_LEN,
        "time_freq": TIME_FREQ,
        "random_seed": RANDOM_SEED,
        "scaler_x": scaler_x,
        "scaler_y": scaler_y,
        "scaler_x_mean": np.asarray(scaler_x.mean_, dtype=np.float64),
        "scaler_x_scale": np.asarray(scaler_x.scale_, dtype=np.float64),
        "scaler_y_mean": np.asarray(scaler_y.mean_, dtype=np.float64),
        "scaler_y_scale": np.asarray(scaler_y.scale_, dtype=np.float64),
        "power_scale_ratio": float(power_scale_ratio),
        "power_scale_offset": float(power_scale_offset),
        "power_reference_mw": power_reference,
        "power_reference_kind": "train_power_q999",
        # Compatibility for source builders; never described as nameplate.
        "capacity": power_reference,
        "regime_feature_config": regime_config,
        "regime_feature_config_path": str(paths["regime_config"].resolve()),
        "regime_feature_config_sha256": sha256_file(paths["regime_config"]),
        "power_reference_path": str(paths["power_reference"].resolve()),
        "power_reference_sha256": sha256_file(paths["power_reference"]),
        "selected_regime_feature_groups": ["P", "H", "D"],
        "selected_regime_feature_names": list(F7_FEATURE_NAMES),
        "split_boundaries": {
            "common_start": grid[0].isoformat(),
            "train_end": grid[train_stop - 1].isoformat(),
            "validation_start": grid[train_stop].isoformat(),
            "validation_end": grid[val_stop - 1].isoformat(),
            "test_start": grid[val_stop].isoformat(),
            "test_end": grid[-1].isoformat(),
            "train_stop_index": train_stop,
            "validation_stop_index": val_stop,
        },
        "timestamp_rules": {
            "power": asdict(power_rule),
            "weather": asdict(weather_rule),
            "power_available_at_rule": (
                f"raw_timestamp+{power_rule.availability_shift_minutes}min"
            ),
            "weather_available_at_rule": (
                f"raw_timestamp+{weather_rule.availability_shift_minutes}min"
            ),
            "history_eligibility_rule": "available_at<=forecast_origin",
            "target_eligibility_rule": (
                "interval_start>=forecast_origin and "
                "available_at>forecast_origin"
            ),
            "power_wind_relative_availability_offset_min": (
                power_rule.availability_shift_minutes
                - weather_rule.availability_shift_minutes
            ),
            "alignment_status": (
                "uncertain_alignment"
                if "assumed" in power_rule.semantics
                or "assumed" in weather_rule.semantics
                else "resolved"
            ),
        },
        "array_path": str(paths["array"].resolve()),
        "array_sha256": array_sha,
        "canonical_path": str(paths["canonical"].resolve()),
        "canonical_sha256": sha256_file(paths["canonical"]),
        "split_indices_path": str(paths["split_indices"].resolve()),
        "split_indices_sha256": sha256_file(paths["split_indices"]),
        "raw_files": [
            power_file_audit,
            weather_file_audit,
            *run_record_audit,
        ],
        "run_records_used_for_features": False,
        "run_records_used_as_nameplate_capacity": False,
        "preprocessing_code_path": str(Path(__file__).resolve()),
        "preprocessing_code_sha256": sha256_file(__file__),
    }
    joblib.dump(bundle, paths["bundle"], compress=3)
    bundle_sha = sha256_file(paths["bundle"])

    summary = {
        "farm_id": farm_id,
        "status": "complete",
        "rows": n_rows,
        "common_start": grid[0].isoformat(),
        "common_end": grid[-1].isoformat(),
        "train_end": grid[train_stop - 1].isoformat(),
        "validation_start": grid[train_stop].isoformat(),
        "validation_end": grid[val_stop - 1].isoformat(),
        "test_start": grid[val_stop].isoformat(),
        "test_end": grid[-1].isoformat(),
        "train_rows": int(train_stop),
        "validation_rows": int(val_stop - train_stop),
        "test_rows": int(n_rows - val_stop),
        "scaler_fit_points": int(scaler_fit_mask.sum()),
        "power_sampling_minutes": power_rule.sampling_minutes,
        "weather_sampling_minutes": weather_rule.sampling_minutes,
        "power_timestamp_semantics": power_rule.semantics,
        "weather_timestamp_semantics": weather_rule.semantics,
        "power_available_at_shift_minutes": (
            power_rule.availability_shift_minutes
        ),
        "power_available_at_rule": (
            f"raw_timestamp+{power_rule.availability_shift_minutes}min"
        ),
        "weather_available_at_shift_minutes": (
            weather_rule.availability_shift_minutes
        ),
        "power_wind_relative_availability_offset_minutes": (
            power_rule.availability_shift_minutes
            - weather_rule.availability_shift_minutes
        ),
        "power_wind_relative_offset": (
            power_rule.availability_shift_minutes
            - weather_rule.availability_shift_minutes
        ),
        "alignment_status": bundle["timestamp_rules"]["alignment_status"],
        "power_wind_alignment_status": bundle["timestamp_rules"][
            "alignment_status"
        ],
        "run_record_file_count": len(run_record_audit),
        "power_valid_points": int(power_valid.sum()),
        "essential_wind_valid_points": int(essential_wind_valid.sum()),
        "train_windows": int(len(origins["train"])),
        "validation_windows": int(len(origins["val"])),
        "test_windows": int(len(origins["test"])),
        "cross_split_target_overlap_count": int(
            sum(cross_split_target_overlap.values())
        ),
        "train_window_retention_ratio": float(retention),
        "training_feasibility": feasibility,
        "limited_test_coverage": bool(len(origins["test"]) < 5000),
        "insufficient_test_samples": bool(len(origins["test"]) < 2000),
        "power_reference_mw": power_reference,
        "power_reference_kind": "train_power_q999",
        "power_reference_path": str(paths["power_reference"].resolve()),
        "regime_feature_config_path": str(paths["regime_config"].resolve()),
        "schema_hash": schema_hash,
        "input_channels_available": input_channels_available,
        "missing_channels": json.dumps(missing_channels, ensure_ascii=False),
        "final_model_input_channels": len(FEATURE_SCHEMA),
        "structurally_missing_speed_channels": json.dumps(
            [
                name
                for name, info in speed_audit["reliability"].items()
                if not info["reliable_train"]
            ],
            ensure_ascii=False,
        ),
        "structurally_missing_direction_channels": json.dumps(
            direction_audit["structurally_missing"], ensure_ascii=False
        ),
        "direction_unknown_rate_per_mille": direction_audit[
            "unknown_rate_per_mille"
        ],
        "direction_unity_violations_after_repair_per_mille": direction_audit[
            "unity_violations_after_repair_per_mille"
        ],
        "power_duplicate_rows": duplicate_power["duplicate_rows"],
        "weather_duplicate_rows": duplicate_weather["duplicate_rows"],
        "array_path": str(paths["array"].resolve()),
        "array_sha256": array_sha,
        "bundle_path": str(paths["bundle"].resolve()),
        "bundle_sha256": bundle_sha,
        "split_indices_path": str(paths["split_indices"].resolve()),
        "split_indices_sha256": sha256_file(paths["split_indices"]),
        "elapsed_seconds": time.monotonic() - started,
    }
    manifest = {
        "status": "complete",
        "created_at": utc_now(),
        "protocol_version": PROTOCOL_VERSION,
        "summary": summary,
        "power_file_audit": power_file_audit,
        "weather_file_audit": weather_file_audit,
        "run_record_file_audit": run_record_audit,
        "duplicate_power": duplicate_power,
        "duplicate_weather": duplicate_weather,
        "power_cleaning": power_clean_audit,
        "weather_cleaning": weather_clean_audit,
        "speed_audit": speed_audit,
        "direction_audit": direction_audit,
        "meteorology_audit": met_audit,
        "cross_split_target_overlap": cross_split_target_overlap,
    }
    atomic_json(paths["farm_manifest"], manifest)
    print(
        f"[complete] {farm_id}: train={len(origins['train']):,}, "
        f"val={len(origins['val']):,}, test={len(origins['test']):,}, "
        f"tr-ref={power_reference:.4f} MW"
    )
    return summary


def ensure_layout() -> None:
    subdirs = (
        "data_audit",
        "prepared_data/feature_arrays",
        "prepared_data/canonical_15min",
        "prepared_data/split_indices",
        "preprocess",
        "manifests/preprocess",
        "models",
        "weights",
        "history",
        "tensorboard",
        "validation_metrics",
        "complexity",
        "visualizations/data_quality",
        "visualizations/training",
        "visualizations/predictions",
        "visualizations/aggregate",
        "testdata_predict_output",
        "partial_runs",
        "attempts",
    )
    for subdir in subdirs:
        (RESULT_ROOT / subdir).mkdir(parents=True, exist_ok=True)
    schema_hash = sha256_json(list(FEATURE_SCHEMA))
    if schema_hash != LEGACY_F7_SCHEMA_HASH:
        raise AssertionError(
            "FEATURE_SCHEMA_V1与已核验的原五场站F7 input_cols发生漂移"
        )
    schema_payload = {
        "schema_version": SCHEMA_VERSION,
        "protocol_version": PROTOCOL_VERSION,
        "input_cols": list(FEATURE_SCHEMA),
        "input_dim": len(FEATURE_SCHEMA),
        "target_index": 44,
        "schema_hash": schema_hash,
        "legacy_f7_schema_hash": LEGACY_F7_SCHEMA_HASH,
        "semantic_mapping_status": "exact_legacy_f7_schema_match",
    }
    atomic_json(RESULT_ROOT / "manifests" / "feature_schema_v1.json", schema_payload)
    power_protocol = {
        "protocol_version": PROTOCOL_VERSION,
        "primary_reference_kind": "train_power_q999",
        "quantile": POWER_REFERENCE_QUANTILE,
        "quantile_method": "linear",
        "fit_scope": "unique_valid_training_timestamps_only",
        "validation_or_test_used": False,
        "used_as_prediction_upper_clip": False,
        "metric_names": ["trNMAE", "trNRMSE"],
    }
    atomic_json(
        RESULT_ROOT / "data_audit" / "round3_power_reference_protocol.json",
        power_protocol,
    )


def parse_farms(raw: str | None) -> list[str]:
    if not raw or raw.strip().lower() in {"all", "*"}:
        return list(EXPECTED_FARMS)
    farms = [item.strip().upper() for item in raw.split(",") if item.strip()]
    invalid = sorted(set(farms) - set(EXPECTED_FARMS))
    if invalid:
        raise ValueError(f"未知场站: {invalid}")
    return list(dict.fromkeys(farms))


def save_data_audit_visualization(frame: pd.DataFrame) -> Path:
    """Save a compact, paper-traceable overview of the prepared stations."""
    os.environ.setdefault(
        "MPLCONFIGDIR",
        str((RESULT_ROOT / "matplotlib_cache").resolve()),
    )
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    ordered = frame.sort_values("farm_id")
    x = np.arange(len(ordered))
    fig, axes = plt.subplots(3, 1, figsize=(15, 12), sharex=True)
    axes[0].bar(
        x,
        ordered["train_windows"],
        label="train",
        color="tab:blue",
    )
    axes[0].bar(
        x,
        ordered["validation_windows"],
        bottom=ordered["train_windows"],
        label="validation",
        color="tab:orange",
    )
    axes[0].bar(
        x,
        ordered["test_windows"],
        bottom=ordered["train_windows"] + ordered["validation_windows"],
        label="test",
        color="tab:green",
    )
    axes[0].set_ylabel("Eligible windows")
    axes[0].set_title("Leakage-free 96-to-16 windows by split")
    axes[0].legend(ncol=3)

    axes[1].bar(
        x,
        ordered["train_window_retention_ratio"],
        color="tab:purple",
    )
    axes[1].axhline(0.8, color="black", linestyle="--", linewidth=1)
    axes[1].set_ylim(0.0, 1.05)
    axes[1].set_ylabel("Train retention")
    axes[1].set_title("Eligible-window retention after causal validity checks")

    axes[2].bar(
        x,
        ordered["power_reference_mw"],
        color="tab:red",
    )
    axes[2].set_ylabel("MW")
    axes[2].set_title("Train-only Q99.9 power reference")
    axes[2].set_xticks(x)
    axes[2].set_xticklabels(ordered["farm_id"], rotation=45, ha="right")
    for axis in axes:
        axis.grid(axis="y", alpha=0.25)
    fig.tight_layout()
    path = (
        RESULT_ROOT
        / "visualizations"
        / "data_quality"
        / "round3_external14_preprocess_overview.png"
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)
    return path


def write_global_outputs(summaries: list[dict[str, Any]], requested: list[str]) -> None:
    frame = pd.DataFrame(summaries).sort_values("farm_id")
    audit_path = RESULT_ROOT / "data_audit" / "round3_raw_data_audit.csv"
    frame.to_csv(audit_path, index=False, encoding="utf-8-sig")
    frame.to_csv(
        RESULT_ROOT / "data_audit" / "round3_external14_data_audit.csv",
        index=False,
        encoding="utf-8-sig",
    )
    split_cols = [
        "farm_id",
        "common_start",
        "train_end",
        "validation_start",
        "validation_end",
        "test_start",
        "test_end",
        "train_rows",
        "validation_rows",
        "test_rows",
        "train_windows",
        "validation_windows",
        "test_windows",
        "cross_split_target_overlap_count",
    ]
    frame[split_cols].to_csv(
        RESULT_ROOT / "data_audit" / "round3_external14_split_manifest.csv",
        index=False,
        encoding="utf-8-sig",
    )
    feasibility_cols = [
        "farm_id",
        "train_windows",
        "train_window_retention_ratio",
        "training_feasibility",
        "validation_windows",
        "test_windows",
        "limited_test_coverage",
        "insufficient_test_samples",
    ]
    frame[feasibility_cols].to_csv(
        RESULT_ROOT / "data_audit" / "round3_training_feasibility.csv",
        index=False,
        encoding="utf-8-sig",
    )
    reference_cols = [
        "farm_id",
        "power_reference_kind",
        "power_reference_mw",
        "power_valid_points",
        "common_start",
        "common_end",
    ]
    frame[reference_cols].to_csv(
        RESULT_ROOT / "data_audit" / "round3_power_reference_table.csv",
        index=False,
        encoding="utf-8-sig",
    )
    timestamp_cols = [
        "farm_id",
        "power_sampling_minutes",
        "weather_sampling_minutes",
        "power_timestamp_semantics",
        "weather_timestamp_semantics",
        "power_available_at_shift_minutes",
        "power_available_at_rule",
        "weather_available_at_shift_minutes",
        "power_wind_relative_availability_offset_minutes",
        "power_wind_relative_offset",
        "alignment_status",
        "power_wind_alignment_status",
        "common_start",
        "common_end",
    ]
    frame[timestamp_cols].to_csv(
        RESULT_ROOT / "data_audit" / "round3_timestamp_semantics.csv",
        index=False,
        encoding="utf-8-sig",
    )
    regime_cols = [
        "farm_id",
        "schema_hash",
        "input_channels_available",
        "missing_channels",
        "final_model_input_channels",
        "regime_feature_config_path",
        "structurally_missing_speed_channels",
        "structurally_missing_direction_channels",
        "direction_unity_violations_after_repair_per_mille",
    ]
    regime_validation = frame[regime_cols].copy()
    regime_validation["regime_config_status"] = np.where(
        (regime_validation["final_model_input_channels"] == 45)
        & (
            regime_validation[
                "direction_unity_violations_after_repair_per_mille"
            ]
            == 0
        ),
        "valid",
        "blocked",
    )
    regime_validation.to_csv(
        RESULT_ROOT / "data_audit" / "round3_regime_config_validation.csv",
        index=False,
        encoding="utf-8-sig",
    )
    overview_path = save_data_audit_visualization(frame)
    complete = set(requested) == set(EXPECTED_FARMS) and set(frame["farm_id"]) == set(
        EXPECTED_FARMS
    )
    payload = {
        "status": "complete" if complete else "partial",
        "created_at": utc_now(),
        "protocol_version": PROTOCOL_VERSION,
        "expected_farms": list(EXPECTED_FARMS),
        "requested_farms": requested,
        "completed_farms": frame["farm_id"].tolist(),
        "farm_count": len(frame),
        "feature_schema_hash": sha256_json(list(FEATURE_SCHEMA)),
        "audit_path": str(audit_path.resolve()),
        "audit_sha256": sha256_file(audit_path),
        "data_overview_path": str(overview_path.resolve()),
        "data_overview_sha256": sha256_file(overview_path),
        "farm_artifacts": [
            {
                "farm_id": row["farm_id"],
                "array_path": row["array_path"],
                "array_sha256": row["array_sha256"],
                "bundle_path": row["bundle_path"],
                "bundle_sha256": row["bundle_sha256"],
                "split_indices_path": row["split_indices_path"],
                "split_indices_sha256": row["split_indices_sha256"],
            }
            for row in frame.to_dict(orient="records")
        ],
    }
    marker = (
        RESULT_ROOT / "round3_preprocess_bundle_complete.json"
        if complete
        else RESULT_ROOT / "partial_runs" / "round3_preprocess_partial.json"
    )
    atomic_json(marker, payload)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Part-3 Round-3 JSFD14 leakage-free preprocessing"
    )
    parser.add_argument("--farms", default="all", help="all或逗号分隔JSFD编号")
    parser.add_argument("--force", action="store_true", help="覆盖已有场站产物")
    parser.add_argument(
        "--timestamp-semantics",
        help="可选权威时间戳语义JSON；未提供时使用保守起点标签假设",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    raw_resolved = RAW_ROOT.resolve()
    if not raw_resolved.exists():
        raise FileNotFoundError(raw_resolved)
    if any(part.lower() == "processed_npz" for part in raw_resolved.parts):
        raise ValueError(f"RAW_ROOT不得指向processed_npz或其子目录: {raw_resolved}")
    ensure_layout()
    farms = parse_farms(args.farms)
    semantics = load_semantics_config(args.timestamp_semantics)
    summaries = [
        prepare_farm(farm_id, semantics, force=bool(args.force)) for farm_id in farms
    ]
    write_global_outputs(summaries, farms)
    print(
        f"Round-3预处理完成: {len(summaries)}个场站；"
        f"结果目录={RESULT_ROOT.resolve()}"
    )


if __name__ == "__main__":
    main()
