"""Prepare JSFD supplementary wind farms for tuned PatchTST teacher pretraining.

This module intentionally has no TensorFlow dependency.  The project training
environment does not currently contain an Excel reader, so run this script in
an environment with ``openpyxl`` once, then let the deep-learning environment
consume the generated, version-independent ``.npz`` caches.
"""

import argparse
import json
import math
import os
from pathlib import Path

import numpy as np
import pandas as pd


TIME_FREQ = '15min'
HISTORY_LEN = 96
FORECAST_LEN = 16
DEFAULT_SOURCE_DIR = Path('./wind_split/supplementary_other_wind_data')
DEFAULT_CACHE_DIR = DEFAULT_SOURCE_DIR / 'processed_npz'
DEFAULT_REPORT_PATH = DEFAULT_CACHE_DIR / 'supplementary_preprocess_report.csv'

WIND_SPEED_COLS = [
    '10米风速',
    '30米风速',
    '50米风速',
    '70米风速',
    '轮毂高度风速',
]
WIND_DIR_COLS = [
    '10米风向',
    '30米风向',
    '50米风向',
    '70米风向',
    '轮毂高度风向',
]
WEATHER_COLS = WIND_SPEED_COLS + WIND_DIR_COLS + [
    '10m气温',
    '10m气压',
    '10m湿度',
]

POWER_COLUMN_ALIASES = {
    '时间': {'时间'},
    '功率': {'实际功率(MW)', '功率-MW', '功率（MW）', '功率'},
}
WEATHER_COLUMN_ALIASES = {
    '时间': {'时间'},
    '10米风速': {'10米高度处风速（m/s）', '10米高度处风速(m/s)', '10米风速'},
    '10米风向': {'10米高度处风向（°）', '10米高度处风向(°)', '10米风向'},
    '30米风速': {'30米高度处风速（m/s）', '30米高度处风速(m/s)', '30米风速'},
    '30米风向': {'30米高度处风向（°）', '30米高度处风向(°)', '30米风向'},
    '50米风速': {'50米高度处风速（m/s）', '50米高度处风速(m/s)', '50米风速'},
    '50米风向': {'50米高度处风向（°）', '50米高度处风向(°)', '50米风向'},
    '70米风速': {'70米高度处风速（m/s）', '70米高度处风速(m/s)', '70米风速'},
    '70米风向': {'70米高度处风向（°）', '70米高度处风向(°)', '70米风向'},
    '轮毂高度风速': {
        '风机轮毂高度处风速（m/s）',
        '风机轮毂高度处风速(m/s)',
        '轮毂高度风速',
    },
    '轮毂高度风向': {
        '风机轮毂高度处风向（°）',
        '风机轮毂高度处风向(°)',
        '轮毂高度风向',
    },
    '10m气温': {'气温（°C）', '气温(°C)', '10m气温'},
    '10m气压': {'气压（hpa）', '气压(hpa)', '10m气压'},
    '10m湿度': {'相对湿度（%）', '相对湿度(%)', '10m湿度'},
}
OPERATION_COLUMN_ALIASES = {
    '开始': {'起始日期', '开始时间'},
    '结束': {'终止时间', '结束时间'},
    '上限': {'最大出力上限值(MW)', '最大出力上限值（MW）'},
}

PHYSICAL_DEFAULTS = {
    '10m气温': 15.0,
    '10m气压': 1000.0,
    '10m湿度': 60.0,
}
WEATHER_LIMITS = {
    '10m气温': (-50.0, 60.0),
    '10m气压': (850.0, 1100.0),
    '10m湿度': (0.0, 100.0),
}


def _normalize_header(value):
    return str(value).replace('\n', '').strip() if value is not None else ''


def _alias_lookup(column_aliases):
    lookup = {}
    for canonical, aliases in column_aliases.items():
        for alias in aliases | {canonical}:
            lookup[_normalize_header(alias)] = canonical
    return lookup


def _parse_timestamp(value):
    if value is None:
        return pd.NaT
    try:
        return pd.to_datetime(value, errors='coerce')
    except (TypeError, ValueError, OverflowError):
        return pd.NaT


def read_excel_records(path, column_aliases, required_columns,
                       blank_stop=512):
    """Read useful rows only, stopping before Excel's formatted empty tail."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            '读取补充数据需要openpyxl。请先在预处理环境安装openpyxl，'
            '或使用项目当前base Python执行本脚本。'
        ) from exc

    lookup = _alias_lookup(column_aliases)
    records = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue

            index_by_column = {}
            for index, value in enumerate(header):
                canonical = lookup.get(_normalize_header(value))
                if canonical and canonical not in index_by_column:
                    index_by_column[canonical] = index
            if not set(required_columns).issubset(index_by_column):
                continue

            seen_valid_row = False
            consecutive_blank_rows = 0
            for row in rows:
                timestamp_index = index_by_column.get(
                    '时间',
                    index_by_column.get('开始'),
                )
                timestamp = _parse_timestamp(
                    row[timestamp_index]
                    if timestamp_index is not None and timestamp_index < len(row)
                    else None
                )
                if pd.isna(timestamp):
                    if seen_valid_row:
                        consecutive_blank_rows += 1
                        if consecutive_blank_rows >= blank_stop:
                            break
                    continue

                seen_valid_row = True
                consecutive_blank_rows = 0
                record = {}
                for canonical, column_index in index_by_column.items():
                    record[canonical] = (
                        row[column_index]
                        if column_index < len(row)
                        else None
                    )
                records.append(record)
    finally:
        workbook.close()

    if not records:
        return pd.DataFrame(columns=list(column_aliases))

    frame = pd.DataFrame.from_records(records)
    for column in frame.columns:
        if column in {'时间', '开始', '结束'}:
            frame[column] = pd.to_datetime(frame[column], errors='coerce')
        else:
            frame[column] = pd.to_numeric(frame[column], errors='coerce')
    timestamp_column = '时间' if '时间' in frame.columns else '开始'
    frame = frame.dropna(subset=[timestamp_column])
    frame = frame.sort_values(timestamp_column)
    frame = frame.drop_duplicates(timestamp_column, keep='last')
    return frame.reset_index(drop=True)


def read_operation_records(path):
    if path is None or not path.exists():
        return pd.DataFrame(columns=['开始', '结束', '上限'])

    records = read_excel_records(
        path,
        OPERATION_COLUMN_ALIASES,
        required_columns=['开始', '结束'],
    )
    if records.empty:
        return records
    records['结束'] = pd.to_datetime(records['结束'], errors='coerce')
    records = records.dropna(subset=['开始', '结束'])
    records = records[records['结束'] >= records['开始']]
    return records.reset_index(drop=True)


def circular_resample(series, frequency=TIME_FREQ):
    radians = np.deg2rad(series.astype(float) % 360.0)
    sin_mean = np.sin(radians).resample(frequency).mean()
    cos_mean = np.cos(radians).resample(frequency).mean()
    direction = np.rad2deg(np.arctan2(sin_mean, cos_mean)) % 360.0
    direction[(sin_mean.isna()) & (cos_mean.isna())] = np.nan
    return direction


def resample_weather(weather):
    weather = weather.set_index('时间').sort_index()
    pieces = {}
    for column in WEATHER_COLS:
        if column not in weather:
            pieces[column] = pd.Series(dtype=float)
        elif column in WIND_DIR_COLS:
            pieces[column] = circular_resample(weather[column])
        else:
            pieces[column] = weather[column].resample(TIME_FREQ).mean()
    return pd.DataFrame(pieces)


def sanitize_weather(weather):
    weather = weather.copy()
    for column in WIND_SPEED_COLS:
        weather[column] = pd.to_numeric(
            weather.get(column),
            errors='coerce',
        )
        weather[column] = weather[column].where(
            weather[column].between(0.0, 60.0)
        )
    for column in WIND_DIR_COLS:
        weather[column] = pd.to_numeric(
            weather.get(column),
            errors='coerce',
        )
        weather[column] = weather[column] % 360.0
    for column, (lower, upper) in WEATHER_LIMITS.items():
        weather[column] = pd.to_numeric(
            weather.get(column),
            errors='coerce',
        )
        weather[column] = weather[column].where(
            weather[column].between(lower, upper)
        )
    return weather


def _circular_row_mean(frame):
    radians = np.deg2rad(frame.to_numpy(dtype=float))
    valid_count = np.sum(np.isfinite(radians), axis=1)
    sin_mean = np.divide(
        np.nansum(np.sin(radians), axis=1),
        valid_count,
        out=np.full(len(frame), np.nan, dtype=float),
        where=valid_count > 0,
    )
    cos_mean = np.divide(
        np.nansum(np.cos(radians), axis=1),
        valid_count,
        out=np.full(len(frame), np.nan, dtype=float),
        where=valid_count > 0,
    )
    angle = np.rad2deg(np.arctan2(sin_mean, cos_mean)) % 360.0
    angle[~np.isfinite(sin_mean) | ~np.isfinite(cos_mean)] = np.nan
    return pd.Series(angle, index=frame.index)


def _circular_series_mean(series):
    values = pd.to_numeric(series, errors='coerce').dropna().to_numpy()
    if len(values) == 0:
        return np.nan
    radians = np.deg2rad(values % 360.0)
    return float(
        np.rad2deg(np.arctan2(
            np.mean(np.sin(radians)),
            np.mean(np.cos(radians)),
        )) % 360.0
    )


def impute_weather(weather):
    """Short-gap interpolation plus physically neutral cross-height fills."""
    weather = sanitize_weather(weather)

    for column in WIND_SPEED_COLS + list(WEATHER_LIMITS):
        weather[column] = weather[column].interpolate(
            method='time',
            limit=4,
            limit_direction='both',
        )

    speed_row_median = weather[WIND_SPEED_COLS].median(axis=1, skipna=True)
    for column in WIND_SPEED_COLS:
        weather[column] = weather[column].fillna(speed_row_median)
        station_median = weather[column].median(skipna=True)
        fallback = 0.0 if not np.isfinite(station_median) else station_median
        weather[column] = weather[column].fillna(fallback)

    direction_row_mean = _circular_row_mean(weather[WIND_DIR_COLS])
    for column in WIND_DIR_COLS:
        weather[column] = weather[column].fillna(direction_row_mean)
        station_direction = _circular_series_mean(weather[column])
        fallback = (
            0.0
            if not np.isfinite(station_direction)
            else float(station_direction)
        )
        weather[column] = weather[column].fillna(fallback) % 360.0

    for column, default in PHYSICAL_DEFAULTS.items():
        station_median = weather[column].median(skipna=True)
        fallback = default if not np.isfinite(station_median) else station_median
        weather[column] = weather[column].fillna(fallback)

    return weather.astype(np.float32)


def estimate_capacity(power, operation_records=None):
    positive = pd.to_numeric(power, errors='coerce').clip(lower=0).dropna()
    if positive.empty:
        raise ValueError('补充场站没有可用于估算容量的功率数据')

    candidates = [
        float(positive.quantile(0.999)),
        float(positive.quantile(0.99) * 1.01),
    ]
    if operation_records is not None and '上限' in operation_records:
        limits = pd.to_numeric(
            operation_records['上限'],
            errors='coerce',
        )
        limits = limits[limits > 0]
        if not limits.empty:
            candidates.append(float(limits.quantile(0.999)))
    capacity = max(candidates)
    return max(1.0, float(math.ceil(capacity)))


def operation_restriction_mask(index, operation_records):
    restricted = np.zeros(len(index), dtype=bool)
    if operation_records is None or operation_records.empty:
        return restricted

    index_values = index.to_numpy(dtype='datetime64[ns]')
    for record in operation_records.itertuples(index=False):
        start = np.datetime64(record.开始)
        end = np.datetime64(record.结束)
        left = int(np.searchsorted(index_values, start, side='left'))
        right = int(np.searchsorted(index_values, end, side='right'))
        if left < right:
            restricted[left:right] = True
    return restricted


def count_valid_windows(mask, window_len=HISTORY_LEN + FORECAST_LEN):
    mask = np.asarray(mask, dtype=np.int16)
    if len(mask) < window_len:
        return 0
    sums = np.convolve(mask, np.ones(window_len, dtype=np.int16), mode='valid')
    return int(np.sum(sums == window_len))


def prepare_station(station_dir):
    station_dir = Path(station_dir)
    power_paths = sorted(station_dir.glob('*场站出力*.xlsx'))
    weather_paths = sorted(station_dir.glob('*测风*.xlsx'))
    operation_paths = sorted(station_dir.glob('*运行记录*.xlsx'))
    if len(power_paths) != 1 or len(weather_paths) != 1:
        raise FileNotFoundError(
            f'{station_dir.name} 应各有一个场站出力和测风Excel，'
            f'当前为 {len(power_paths)}/{len(weather_paths)}'
        )

    power = read_excel_records(
        power_paths[0],
        POWER_COLUMN_ALIASES,
        required_columns=['时间', '功率'],
    )
    weather = read_excel_records(
        weather_paths[0],
        WEATHER_COLUMN_ALIASES,
        required_columns=['时间'],
    )
    operations = read_operation_records(
        operation_paths[0] if operation_paths else None
    )
    if power.empty or weather.empty:
        raise ValueError(f'{station_dir.name} 没有有效功率或测风时间行')

    for column in WEATHER_COLS:
        if column not in weather:
            weather[column] = np.nan
    power_15m = (
        power.set_index('时间')['功率']
        .sort_index()
        .resample(TIME_FREQ)
        .mean()
    )
    weather_15m = resample_weather(weather)

    start = max(power_15m.index.min(), weather_15m.index.min())
    end = min(power_15m.index.max(), weather_15m.index.max())
    if start >= end:
        raise ValueError(f'{station_dir.name} 的功率和气象没有重叠时间')
    index = pd.date_range(start.ceil(TIME_FREQ), end.floor(TIME_FREQ), freq=TIME_FREQ)
    power_raw = power_15m.reindex(index)
    weather_raw = sanitize_weather(weather_15m.reindex(index))

    capacity = estimate_capacity(power_raw, operations)
    finite_power = power_raw.notna()
    plausible_power = power_raw.between(-0.10 * capacity, 1.20 * capacity)
    valid_speed = weather_raw[WIND_SPEED_COLS].notna().any(axis=1)
    all_speed_zero = (
        weather_raw[WIND_SPEED_COLS].fillna(0.0).abs().max(axis=1) <= 0.05
    )
    inconsistent_calm = all_speed_zero & (power_raw > 0.02 * capacity)
    restricted = operation_restriction_mask(index, operations)
    quality_mask = (
        finite_power.to_numpy()
        & plausible_power.fillna(False).to_numpy()
        & valid_speed.to_numpy()
        & ~inconsistent_calm.fillna(True).to_numpy()
        & ~restricted
    )

    weather_filled = impute_weather(weather_raw)
    power_clean = power_raw.clip(lower=0.0, upper=capacity)
    power_clean = power_clean.interpolate(
        method='time',
        limit=4,
        limit_direction='both',
    )
    power_clean = power_clean.fillna(0.0).astype(np.float32)

    raw_values = weather_filled[WEATHER_COLS].to_numpy(dtype=np.float32)
    valid_windows = count_valid_windows(quality_mask)
    metadata = {
        'schema_version': 1,
        'station_id': station_dir.name,
        'time_freq': TIME_FREQ,
        'start_time': index.min().isoformat(),
        'end_time': index.max().isoformat(),
        'rows': int(len(index)),
        'estimated_capacity_mw': capacity,
        'quality_points': int(quality_mask.sum()),
        'quality_ratio': float(quality_mask.mean()),
        'valid_windows_112': valid_windows,
        'operation_excluded_points': int(restricted.sum()),
        'power_negative_ratio_before_clip': float(
            (power_raw[finite_power] < 0).mean()
        ),
        'power_zero_ratio_before_clip': float(
            (power_raw[finite_power] == 0).mean()
        ),
        'power_file': str(power_paths[0]),
        'weather_file': str(weather_paths[0]),
        'operation_file': str(operation_paths[0]) if operation_paths else None,
        'raw_feature_columns': WEATHER_COLS,
    }
    return {
        'timestamps_ns': index.to_numpy(dtype='datetime64[ns]').astype(np.int64),
        'raw_values': raw_values,
        'power_mw': power_clean.to_numpy(dtype=np.float32),
        'quality_mask': quality_mask.astype(bool),
        'capacity_mw': np.asarray(capacity, dtype=np.float32),
        'metadata_json': np.asarray(
            json.dumps(metadata, ensure_ascii=False)
        ),
    }, metadata


def save_station_cache(cache_path, arrays):
    cache_path = Path(cache_path)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = cache_path.with_suffix(cache_path.suffix + '.tmp.npz')
    np.savez_compressed(temporary_path, **arrays)
    os.replace(temporary_path, cache_path)


def preprocess_all(source_dir=DEFAULT_SOURCE_DIR,
                   cache_dir=DEFAULT_CACHE_DIR,
                   report_path=DEFAULT_REPORT_PATH,
                   station_names=None,
                   force=False):
    source_dir = Path(source_dir)
    cache_dir = Path(cache_dir)
    report_path = Path(report_path)
    station_filter = set(station_names or [])
    station_dirs = [
        path for path in sorted(source_dir.glob('JSFD*'))
        if path.is_dir() and (
            not station_filter or path.name in station_filter
        )
    ]
    if not station_dirs:
        raise FileNotFoundError(f'未在 {source_dir} 找到JSFD补充场站')

    rows = []
    for station_dir in station_dirs:
        cache_path = cache_dir / f'{station_dir.name}_15min.npz'
        if cache_path.exists() and not force:
            with np.load(cache_path, allow_pickle=False) as cached:
                metadata = json.loads(str(cached['metadata_json'].item()))
            metadata['cache_status'] = 'reused'
            metadata['cache_path'] = str(cache_path)
            rows.append(metadata)
            print(f'{station_dir.name}: 复用 {cache_path}')
            continue

        print(f'{station_dir.name}: 读取并清洗Excel...')
        arrays, metadata = prepare_station(station_dir)
        save_station_cache(cache_path, arrays)
        metadata['cache_status'] = 'created'
        metadata['cache_path'] = str(cache_path)
        rows.append(metadata)
        print(
            f"{station_dir.name}: rows={metadata['rows']}, "
            f"quality={metadata['quality_ratio']:.3f}, "
            f"valid_windows={metadata['valid_windows_112']}, "
            f"capacity≈{metadata['estimated_capacity_mw']:.1f}MW"
        )

    report = pd.DataFrame(rows)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report.to_csv(report_path, index=False, encoding='utf-8-sig')
    print(f'补充数据预处理报告: {report_path}')
    return report


def parse_args():
    parser = argparse.ArgumentParser(
        description='预处理JSFD001-JSFD014补充风电数据'
    )
    parser.add_argument('--source-dir', default=str(DEFAULT_SOURCE_DIR))
    parser.add_argument('--cache-dir', default=str(DEFAULT_CACHE_DIR))
    parser.add_argument('--report-path', default=str(DEFAULT_REPORT_PATH))
    parser.add_argument(
        '--stations',
        default='',
        help='逗号分隔的场站名；留空处理全部',
    )
    parser.add_argument('--force', action='store_true')
    return parser.parse_args()


def main():
    args = parse_args()
    station_names = [
        value.strip()
        for value in args.stations.split(',')
        if value.strip()
    ]
    preprocess_all(
        source_dir=args.source_dir,
        cache_dir=args.cache_dir,
        report_path=args.report_path,
        station_names=station_names,
        force=args.force,
    )


if __name__ == '__main__':
    main()
